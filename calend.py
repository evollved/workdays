import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintOptions
from datetime import datetime, timedelta, date

def create_calendar(start_date, work_days, rest_days):
    # Создаем новую рабочую книгу и активный лист
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Календарь"

    # Устанавливаем альбомную ориентацию и границы
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(left=0.8/2.54, right=0.8/2.54, top=0.8/2.54, bottom=0.8/2.54)
    ws.print_options = PrintOptions(horizontalCentered=True, verticalCentered=True)

    # Названия месяцев и дней недели
    months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
    days_of_week = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

    # Устанавливаем начальную позицию
    start_row = 1
    start_col = 1
    col_width_cm = 2.4571  # Ширина колонок 1=0,35см
    col_spacing_cm = 1.4285  # Ширина промежуточного столбца в сантиметрах
    row_height = 0.64 * 28.3465  # Высота строк (Excel измеряет высоту в пунктах)

    # Функция для определения рабочих дней
    def is_work_day(current_date, start_date, work_days, rest_days):
        delta = (current_date - start_date).days
        cycle_length = work_days + rest_days
        return (delta % cycle_length) < work_days

    # Определение стиля границ
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Генерируем календарь по месяцам
    for month_idx, month in enumerate(months):
        month_start = date(start_date.year, month_idx + 1, 1)
        month_end = (date(start_date.year, month_idx + 2, 1) - timedelta(days=1)) if month_idx < 11 else date(start_date.year, 12, 31)
        month_days = (month_end - month_start).days + 1

        # Устанавливаем название месяца
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 6)
        cell = ws.cell(row=start_row, column=start_col)
        cell.value = month
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

        # Устанавливаем дни недели
        for col_idx, day in enumerate(days_of_week):
            cell = ws.cell(row=start_row + 1, column=start_col + col_idx)
            cell.value = day
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Заполняем дни месяца
        current_row = start_row + 2
        current_col = start_col

        # Определяем, на какой день недели начинается месяц
        first_day_of_month = month_start.weekday()  # 0 - понедельник, 6 - воскресенье

        # Заполняем пустые ячейки до первого дня месяца
        for _ in range(first_day_of_month):
            cell = ws.cell(row=current_row, column=current_col)
            cell.value = ""
            cell.border = thin_border
            current_col += 1

        for day in range(1, month_days + 1):
            current_date = date(start_date.year, month_idx + 1, day)
            is_work = is_work_day(current_date, start_date, work_days, rest_days)

            cell = ws.cell(row=current_row, column=current_col)
            cell.value = day
            cell.font = Font(name='Arial', size=16)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if is_work:
                cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

            # Переход на следующий день
            current_col += 1
            if current_col > start_col + 6:  # Конец недели
                current_col = start_col
                current_row += 1

        # Настраиваем ширину колонок и высоту строк
        for col_idx in range(start_col, start_col + 7):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width_cm * 2.54 * 7.0 / 10.0
        for row_idx in range(start_row + 1, current_row + 1):
            ws.row_dimensions[row_idx].height = row_height

        # Перемещение к следующему месяцу
        if (month_idx + 1) % 4 == 0:
            start_row += current_row - start_row + 2  # Смещение на необходимое количество строк вниз для новой строки месяцев плюс 2 строки
            start_col = 1  # Начало новой строки
        else:
            # Добавляем промежуточный столбец
            spacer_col = start_col + 7
            ws.column_dimensions[get_column_letter(spacer_col)].width = col_spacing_cm * 2.54 * 7.0 / 10.0
            start_col += 8  # Смещение на 8 колонок вправо для следующего месяца

    # Сохраняем книгу
    wb.save("Календарь.xlsx")
    print("Календарь сохранен как 'Календарь.xlsx'")

def main():
    # Запрашиваем у пользователя начальную дату и график работы
    start_date_str = input("Введите дату начала работы (в формате ГГГГ-ММ-ДД): ")
    work_days = int(input("Введите количество рабочих дней подряд: "))
    rest_days = int(input("Введите количество выходных дней подряд: "))

    # Преобразуем введенную дату в объект datetime
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()

    # Генерируем календарь
    create_calendar(start_date, work_days, rest_days)

if __name__ == "__main__":
    main()
